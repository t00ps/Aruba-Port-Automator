import paramiko
from tkinter import *
from tkinter import messagebox, scrolledtext
from docx import Document
import time
from datetime import datetime
import os
import re
from openpyxl import Workbook, load_workbook

# --- MAC address functions ---
def normalize_mac(mac_address):
    cleaned = re.sub(r'[^0-9A-Fa-f]', '', mac_address)
    if len(cleaned) != 12:
        return None
    return ':'.join(cleaned[i:i+2] for i in range(0, 12, 2)).lower()

def validate_mac(mac_address):
    normalized = normalize_mac(mac_address)
    return normalized

def validate_patch_panel(patch_panel):
    pattern = r'^\d{2}-\d{3}-\d{2}$'
    return re.match(pattern, patch_panel)

# --- SSH functions ---
def connect_ssh(ip, username, password):
    try:
        ssh = paramiko.SSHClient()
        ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        print("Nawiązywanie połączenia SSH...")
        ssh.connect(ip, username=username, password=password)
        time.sleep(1)
        shell = ssh.invoke_shell()
        shell.send('\n')
        time.sleep(1)
        print(f"Połączono z {ip}")
        return shell, ssh
    except Exception as e:
        messagebox.showerror("Błąd połączenia", f"Nie udało się połączyć z urządzeniem: {e}")
        return None, None

def show_intrusion_log():
    ip = entry_ip.get()
    username = entry_username.get()
    password = entry_password.get()

    if not ip or not username or not password:
        messagebox.showerror("Błąd", "Proszę wypełnić IP, login i hasło.")
        return

    shell, ssh = connect_ssh(ip, username, password)
    if shell:
        shell.send("show port-security intrusion-log\n")
        time.sleep(2)
        output = shell.recv(5000).decode()

        log_window = Toplevel(root)
        log_window.title("Logi intruzji")
        log_text = scrolledtext.ScrolledText(log_window, wrap=WORD, width=80, height=15)
        log_text.pack(padx=10, pady=10)
        log_text.insert(INSERT, output)
        log_text.configure(state='disabled')

        ssh.close()

def send_command(shell, port, vlan, mac_address):
    try:
        print(f"Konfigurowanie portu {port}, VLAN {vlan}, MAC {mac_address}")
        shell.send("configure terminal\n")
        time.sleep(1)

        shell.send(f"interface {port}\n")
        time.sleep(1)
        shell.send(f"vlan {vlan} untagged {port}\n")
        time.sleep(1)
        shell.send("exit\n")
        time.sleep(1)

        shell.send(f"no port-security {port}\n")
        time.sleep(1)
        shell.send(f"port-security {port} learn-mode configured mac-address {mac_address}\n")
        time.sleep(1)

        shell.send("write memory\n")
        time.sleep(2)

        output = shell.recv(9999).decode()
        print(f"Wynik konfiguracji:\n{output}")
        return output
    except Exception as e:
        print(f"Błąd podczas wysyłania komend: {e}")
        return str(e)

def autosize_columns(ws):
    for column_cells in ws.columns:
        max_length = 0
        column = column_cells[0].column_letter
        for cell in column_cells:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[column].width = max_length + 2

def save_to_excel(ip, port, vlan, mac_address, patch_panel):
    try:
        filename = "autoryzacja_sieciowa.xlsx"
        filepath = os.path.join(os.path.dirname(__file__), filename)

        if os.path.exists(filepath):
            wb = load_workbook(filepath)
        else:
            wb = Workbook()
            default_sheet = wb.active
            if default_sheet and default_sheet.title == "Sheet" and default_sheet.max_row == 1 and default_sheet.max_column == 1 and not default_sheet['A1'].value:
                wb.remove(default_sheet)

        sheet_name = ip
        if sheet_name not in wb.sheetnames:
            ws = wb.create_sheet(title=sheet_name)
            ws['A1'] = "Patch Panel"
            ws['B1'] = "Switch Port"
            ws['C1'] = "MAC (authorized)"
            ws['D1'] = "VLAN"
            ws['E1'] = "Timestamp"
        else:
            ws = wb[sheet_name]
            if ws['A1'].value != "Patch Panel":
                ws.insert_rows(1)
                ws['A1'] = "Patch Panel"
                ws['B1'] = "Switch Port"
                ws['C1'] = "MAC (authorized)"
                ws['D1'] = "VLAN"
                ws['E1'] = "Timestamp"

        try:
            port_num = int(port)
        except:
            messagebox.showerror("Błąd", f"Port musi być liczbą całkowitą. Podano: {port}")
            return

        target_row = port_num + 1

        max_row = ws.max_row
        if max_row < target_row:
            for r in range(max_row + 1, target_row):
                ws.append([None, None, None, None, None])

        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        ws.cell(row=target_row, column=1, value=patch_panel)
        ws.cell(row=target_row, column=2, value=port_num)
        ws.cell(row=target_row, column=3, value=mac_address)
        ws.cell(row=target_row, column=4, value=vlan)
        ws.cell(row=target_row, column=5, value=timestamp)

        autosize_columns(ws)
        wb.save(filepath)

        print(f"Excel zaktualizowany: {os.path.abspath(filepath)}")
        messagebox.showinfo("Zapis", f"Dane zapisane do: {filename} w arkuszu {sheet_name}")

    except Exception as e:
        print(f"Błąd zapisu do Excela: {e}")
        messagebox.showerror("Błąd zapisu", f"Nie udało się zapisać Excela: {e}")

# --- Submit handler ---
def handle_submit():
    ip = entry_ip.get()
    username = entry_username.get()
    password = entry_password.get()
    port = entry_port.get()
    vlan = entry_vlan.get()
    mac_address = entry_mac.get()
    patch_panel = entry_patch_panel.get()

    if not validate_patch_panel(patch_panel):
        messagebox.showerror("Błąd", "Nieprawidłowy format Patch Panel! Użyj formatu: xx-xxx-xx")
        return

    normalized_mac = validate_mac(mac_address)
    if not normalized_mac:
        messagebox.showerror("Błąd", "Nieprawidłowy format adresu MAC!")
        return

    mac_address = normalized_mac

    shell, ssh = connect_ssh(ip, username, password)
    if shell:
        result = send_command(shell, port, vlan, mac_address)
        if result:
            messagebox.showinfo("Wynik", f"Komenda wykonana:\n{result}")
            save_to_excel(ip, port, vlan, mac_address, patch_panel)
        ssh.close()

# --- GUI setup ---
root = Tk()
root.title("Autoryzacja urządzeń Aruba 2930F")

Label(root, text="IP Switcha").grid(row=0, column=0, sticky=W, padx=5, pady=2)
entry_ip = Entry(root)
entry_ip.grid(row=0, column=1, padx=5, pady=2)

Label(root, text="Login").grid(row=1, column=0, sticky=W, padx=5, pady=2)
entry_username = Entry(root)
entry_username.grid(row=1, column=1, padx=5, pady=2)

Label(root, text="Hasło").grid(row=2, column=0, sticky=W, padx=5, pady=2)
entry_password = Entry(root, show="*")
entry_password.grid(row=2, column=1, padx=5, pady=2)

Button(root, text="Pokaż logi intruzji", command=show_intrusion_log).grid(row=3, columnspan=2, pady=8)

Label(root, text="Patch Panel (xx-xxx-xx)").grid(row=4, column=0, sticky=W, padx=5, pady=2)
entry_patch_panel = Entry(root)
entry_patch_panel.grid(row=4, column=1, padx=5, pady=2)

Label(root, text="Port (np. 1)").grid(row=5, column=0, sticky=W, padx=5, pady=2)
entry_port = Entry(root)
entry_port.grid(row=5, column=1, padx=5, pady=2)

Label(root, text="VLAN (np. 20)").grid(row=6, column=0, sticky=W, padx=5, pady=2)
entry_vlan = Entry(root)
entry_vlan.grid(row=6, column=1, padx=5, pady=2)

Label(root, text="Adres MAC").grid(row=7, column=0, sticky=W, padx=5, pady=2)
entry_mac = Entry(root)
entry_mac.grid(row=7, column=1, padx=5, pady=2)

Button(root, text="Zatwierdź", command=handle_submit).grid(row=8, columnspan=2, pady=10)

root.mainloop()
